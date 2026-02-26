#!/usr/bin/env python3
"""Verify customer status from Excel against Stripe subscription data."""

import requests
import urllib3
import openpyxl
import time
from datetime import datetime

urllib3.disable_warnings()

API_KEY = "***REMOVED***"
EXCEL_PATH = "/Users/edoardopelli/projects/Customer Success/New customers since november.xlsx"

def get_stripe_customer(email):
    """Look up a customer in Stripe by email."""
    r = requests.get(
        "https://api.stripe.com/v1/customers",
        params={"email": email, "limit": 1},
        auth=(API_KEY, ""),
    )
    data = r.json()
    if data.get("data"):
        return data["data"][0]
    return None

def get_subscriptions(customer_id):
    """Get all subscriptions for a customer (including canceled)."""
    r = requests.get(
        "https://api.stripe.com/v1/subscriptions",
        params={"customer": customer_id, "limit": 100, "status": "all"},
        auth=(API_KEY, ""),
    )
    return r.json().get("data", [])

def get_latest_charge(customer_id):
    """Get the latest successful charge for a customer."""
    r = requests.get(
        "https://api.stripe.com/v1/charges",
        params={"customer": customer_id, "limit": 1},
        auth=(API_KEY, ""),
    )
    data = r.json().get("data", [])
    if data:
        charge = data[0]
        return {
            "amount": charge["amount"] / 100,
            "currency": charge["currency"].upper(),
            "date": datetime.fromtimestamp(charge["created"]).strftime("%Y-%m-%d"),
            "status": charge["status"],
        }
    return None

def determine_stripe_status(subscriptions):
    """Determine if customer is Active or Churned based on subscriptions."""
    if not subscriptions:
        return "Churned", "No subscriptions found"

    active_statuses = {"active", "trialing"}
    for sub in subscriptions:
        if sub["status"] in active_statuses:
            plan_name = ""
            if sub.get("items", {}).get("data"):
                item = sub["items"]["data"][0]
                plan_name = item.get("price", {}).get("nickname") or item.get("plan", {}).get("nickname") or ""
                if not plan_name:
                    amount = item.get("price", {}).get("unit_amount", 0) / 100
                    currency = item.get("price", {}).get("currency", "").upper()
                    interval = item.get("price", {}).get("recurring", {}).get("interval", "")
                    plan_name = f"{currency} {amount}/{interval}"
            return "Active", f"{sub['status']} - {plan_name}" if plan_name else sub["status"]

    # No active subscription — find the most recent one
    latest = max(subscriptions, key=lambda s: s.get("created", 0))
    canceled_at = ""
    if latest.get("canceled_at"):
        canceled_at = datetime.fromtimestamp(latest["canceled_at"]).strftime("%Y-%m-%d")
    return "Churned", f"{latest['status']} (canceled {canceled_at})" if canceled_at else latest["status"]


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    results = []
    total = ws.max_row - 1  # exclude header

    print(f"Processing {total} customers...\n")

    for row in range(2, ws.max_row + 1):
        cs_manager = ws.cell(row, 1).value
        name = ws.cell(row, 2).value
        email = ws.cell(row, 3).value
        first_sub_date = ws.cell(row, 4).value
        country = ws.cell(row, 5).value
        excel_status = ws.cell(row, 6).value or ""
        odd = ws.cell(row, 7).value

        if not email:
            continue

        idx = row - 1
        print(f"[{idx}/{total}] {name} ({email})...", end=" ", flush=True)

        # Query Stripe
        customer = get_stripe_customer(email)
        if not customer:
            stripe_status = "Not Found"
            detail = "No Stripe customer with this email"
            last_payment = "-"
        else:
            cust_id = customer["id"]
            subs = get_subscriptions(cust_id)
            stripe_status, detail = determine_stripe_status(subs)
            charge = get_latest_charge(cust_id)
            last_payment = charge["date"] if charge else "-"

        # Normalize for comparison
        excel_norm = excel_status.strip().lower() if excel_status else ""
        stripe_norm = stripe_status.strip().lower()
        match = excel_norm == stripe_norm
        discrepancy = "" if match else "MISMATCH"

        results.append({
            "cs_manager": cs_manager,
            "name": name,
            "email": email,
            "country": country,
            "excel_status": excel_status,
            "stripe_status": stripe_status,
            "detail": detail,
            "last_payment": last_payment,
            "discrepancy": discrepancy,
        })

        status_icon = "OK" if match else "MISMATCH"
        print(f"Excel={excel_status} | Stripe={stripe_status} [{status_icon}]")

        # Small delay to respect rate limits
        time.sleep(0.15)

    # Print summary
    print("\n" + "=" * 120)
    print("RESULTS SUMMARY")
    print("=" * 120)

    active_count = sum(1 for r in results if r["stripe_status"] == "Active")
    churned_count = sum(1 for r in results if r["stripe_status"] == "Churned")
    not_found_count = sum(1 for r in results if r["stripe_status"] == "Not Found")
    mismatch_count = sum(1 for r in results if r["discrepancy"] == "MISMATCH")

    print(f"\nTotal customers: {len(results)}")
    print(f"Active (Stripe):    {active_count}")
    print(f"Churned (Stripe):   {churned_count}")
    print(f"Not Found:          {not_found_count}")
    print(f"Discrepancies:      {mismatch_count}")

    # Markdown table
    print("\n\n### Full Customer Status Comparison\n")
    print("| # | CS Mgr | Customer Name | Email | Country | Excel Status | Stripe Status | Subscription Detail | Last Payment | Match |")
    print("|---|--------|---------------|-------|---------|--------------|---------------|---------------------|--------------|-------|")
    for i, r in enumerate(results, 1):
        match_str = "OK" if not r["discrepancy"] else "**MISMATCH**"
        print(f"| {i} | {r['cs_manager']} | {r['name']} | {r['email']} | {r['country']} | {r['excel_status']} | {r['stripe_status']} | {r['detail']} | {r['last_payment']} | {match_str} |")

    # Discrepancies table
    if mismatch_count > 0:
        print(f"\n\n### Discrepancies ({mismatch_count})\n")
        print("| # | Customer Name | Email | Excel Says | Stripe Says | Detail | Last Payment |")
        print("|---|---------------|-------|------------|-------------|--------|--------------|")
        j = 0
        for r in results:
            if r["discrepancy"] == "MISMATCH":
                j += 1
                print(f"| {j} | {r['name']} | {r['email']} | {r['excel_status']} | {r['stripe_status']} | {r['detail']} | {r['last_payment']} |")

    print("\n")

if __name__ == "__main__":
    main()
